import csv
import logging
import sys
import os

import openpyxl # <- pip install openpyxl
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


"""
Program Description:
    Collects the configuration option output csv files in a single, well-formatted excel sheet for further analysis.

Requirements:
    pip install openpyxl

Syntax is:
    python3 collectResults.py <dir_with_info_csv_files>
 
    <dir_with_info_csv_files> needs to contain the following files:
        'buildAccesses.csv', 'buildsOverview.csv' and 'coverage_overview.csv'

    'coverage_overview.csv' is created using the coverageMerger.py

"""


BUILD_ACCESSES_FILE_NAME = "buildAccesses.csv"
BUILD_OVERVIEW_FILE_NAME = "buildsOverview.csv"
BUILD_COVERAGE_FILE_NAME = "coverage_overview.csv"

OUTPUT_FILE_NAME = "results.xlsx"


class ExcelLine:
    def __init__(self, row):
        self.row = row.copy()

    def appendRowToWs(self, ws):
        ws.append(self.row)
        rowIdx = ws.max_row
        cellRow = [ws.cell(row=rowIdx, column=cIdx) for cIdx in range(1,len(self.row)+1)]
        self.configureStandardStyle(cellRow)
        self.configureStyle(cellRow)

    def configureStandardStyle(self, cellRow):
        for cell in cellRow:
            cell.font = Font(name="Consolas")

    # To Overwrite
    def configureStyle(self, cellRow):
        pass


class BuildInfoLine(ExcelLine):
    tag: str
    # Coverage Related
    lC: int = None
    lMax: int = None
    lCov: float = None
    fC: int = None
    fMax: int = None
    fCov: float = None
    # Build Overview Related
    no: int = None
    bTime = None
    optionList: list = []
    # Access Related
    accesses: int = None

    def __init__(self, tag):
        super().__init__([])
        self.tag = tag
        self.bold = False


    def __prepareLine(self):
        self.standardEntries = [self.no, self.tag, self.bTime, self.lC, self.lMax, self.lCov, self.fC, self.fMax, self.fCov, self.accesses]
        self.row = []
        for e in (self.standardEntries+self.optionList):
            if e != None:
                
                self.row.append(e)
            else:
                self.row.append("")

    def appendRowToWs(self, ws):
        self.__prepareLine()
        super().appendRowToWs(ws)

    def configureStyle(self, cellRow):
        cellRow[5].number_format = "0.00%"
        cellRow[8].number_format = "0.00%"
        cellRow[2].alignment = Alignment(horizontal="right")

        sign_check = "Yes"
        sign_cross = "No"
        for cell in cellRow[len(self.standardEntries):]:
            if cell.value == "FLAG_SET":
                cell.value = sign_check
                cell.font = Font(color="76933C", name=cell.font.name) # dark green
            elif  cell.value == "FLAG_NOT_SET":
                cell.value = sign_cross
                cell.font = Font(color="E26B0A", name=cell.font.name) # orange
            cell.alignment = Alignment(horizontal="center")


        if self.bold:
            for cell in cellRow:
                cell.font = Font(name=cell.font.name, bold=True, color=cell.font.color)


class BuildOverviewLine(ExcelLine):
    def __init__(self, row):
        super().__init__(row)
        self.row[0] = int(self.row[0])
        try:
            self.row[2] = float(self.row[2])
            self.existed = False
        except ValueError:
            self.existed = True
        self.no, self.tag, self.bTime = self.row[:3]
        self.optionList = self.row[3:]
        


    def configureStyle(self, cellRow):
        cellRow[2].alignment = Alignment(horizontal="right")

        sign_check = "Yes"
        sign_cross = "No"
        for cell in cellRow[3:]:
            if cell.value == "FLAG_SET":
                cell.value = sign_check
                cell.font = Font(color="76933C", name=cell.font.name) # dark green
            elif  cell.value == "FLAG_NOT_SET":
                cell.value = sign_cross
                cell.font = Font(color="E26B0A", name=cell.font.name) # orange
            cell.alignment = Alignment(horizontal="center")
    

class CoverageInfoLine(ExcelLine):
    def __init__(self, row, bold=False):
        super().__init__(row)
        type = [str, int, int, float, int, int, float]
        for idx, t in enumerate(type):
            self.row[idx] = t(self.row[idx])
        self.tag, self.lines, self.linesMax, self.linesCov, self.func, self.funcMax, self.funcCov = self.row
        
        self.bold = bold
        

    def configureStyle(self, cellRow):
        cellRow[3].number_format = "0.00%"
        cellRow[6].number_format = "0.00%"
        if self.bold:
            for cell in cellRow:
                cell.font = Font(name=cell.font.name, bold=True)


class BuildAccessLine(ExcelLine):
    def __init__(self, row, bold=False):
        super().__init__(row)
        self.row[1] = int(self.row[1])
        self.tag, self.accesses = self.row
        self.bold = bold
        

    def configureStyle(self, cellRow):
        if self.bold:
            for cell in cellRow:
                cell.font = Font(name=cell.font.name, bold=True)


class HeaderLine(ExcelLine):
    def __init__(self, row, alignment):
        super().__init__(row)
        self.alignment = alignment

    def configureStyle(self, cellRow):
        for cell in cellRow:
            cell.font = Font(name=cell.font.name, bold=True)
        for cell,ali in zip(cellRow, self.alignment):
            cell.alignment = Alignment(horizontal=ali)


def adjustColumns(ws):
    dims = {}
    for row in ws:
        for cell in row:
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
    for col, value in dims.items():
        ws.column_dimensions[col].width = max(5, 1.2*value) # Minimum width of 5


def getBuildInfoLineFromTag(buildInfoLines, tag) -> BuildInfoLine:
    if tag in buildInfoLines:
        return buildInfoLines[tag]
    else:
        buildInfoLines[tag] = BuildInfoLine(tag)
        return buildInfoLines[tag]


def appendBuildInfo(ws, buildInfoLines, buildInfoHeader):
    
    buildInfoHeader.appendRowToWs(ws)

    totalLine = None
    if "Total" in buildInfoLines:
        totalLine = buildInfoLines.pop("Total")
        totalLine.bold = True

    lines = [*buildInfoLines.values()]
    if all(x.no!=None for x in lines):
        lines.sort(key=lambda x: x.no)
    for line in lines:
        line.appendRowToWs(ws)

    if totalLine != None:
        totalLine.appendRowToWs(ws)

    ws.auto_filter.ref = "A1:"+get_column_letter(ws.max_column)+str(ws.max_row)
    adjustColumns(ws)


def appendBuildAccessInfo(ws, buildInfoLines, buildInfoHeader):
    accesses_csv_path = os.path.join(path, BUILD_ACCESSES_FILE_NAME)
    
    try:
        with open(accesses_csv_path) as f:
            reader = csv.reader(f, delimiter=',')
            lines = []
            totalLine = None
            for idx,row in enumerate(reader):
                if idx == 0:
                    headerLine = HeaderLine(row, ["left","right"])
                    headerLine.appendRowToWs(ws)
                    buildInfoHeader.row[9] = row[1]
                    buildInfoHeader.row[1] = row[0]

                elif row[0] == "Total":
                    totalLine = BuildAccessLine(row, bold=True)
                    bILine = getBuildInfoLineFromTag(buildInfoLines, "Total")
                    bILine.accesses = bALine.accesses
                else:
                    bALine = BuildAccessLine(row)
                    lines += [bALine]
                    bILine = getBuildInfoLineFromTag(buildInfoLines, bALine.tag)
                    bILine.accesses = bALine.accesses
                    
                
            lines.sort(key=lambda x: x.accesses, reverse=True)
            for l in lines:
                l.appendRowToWs(ws)
            if(totalLine):
                totalLine.appendRowToWs(ws)
            else:
                logging.error("Build Accesses Sheet: Missing 'Total' entry!")

            ws.auto_filter.ref = "A1:"+get_column_letter(ws.max_column)+str(ws.max_row)
            adjustColumns(ws)
    except IOError as error:
        logging.warning(f"Cannot find/open '{ accesses_csv_path}'.\n\nError message: '{error}'\n\nBuild-Accesses-Sheet will be skipped...\n")


def appendCoverageInfo(ws, buildInfoLines, buildInfoHeader):
    cov_csv_path = os.path.join(path, BUILD_COVERAGE_FILE_NAME)
    try:
        with open(cov_csv_path) as f:
            reader = csv.reader(f, delimiter=',')
            for idx,row in enumerate(reader):
                if idx == 0:
                    headerLine = HeaderLine(row, ["left"]+6*["right"])
                    headerLine.appendRowToWs(ws)
                    buildInfoHeader.row[3:9] = row[1:7]
                    buildInfoHeader.row[1] = row[0]
                    buildInfoHeader.alignment[3:9] = 6*["right"]
                elif row[0] == "Collectively":
                    covLine = CoverageInfoLine(row, bold=True)
                    covLine.appendRowToWs(ws)
                    bILine = getBuildInfoLineFromTag(buildInfoLines, "Total")
                    bILine.lC, bILine.lMax, bILine.lCov, bILine.fC, bILine.fMax, bILine.fCov = \
                        covLine.lines, covLine.linesMax, covLine.linesCov, covLine.func, covLine.funcMax, covLine.funcCov
                else:
                    covLine = CoverageInfoLine(row)
                    covLine.appendRowToWs(ws)
                    bILine = getBuildInfoLineFromTag(buildInfoLines, covLine.tag)
                    bILine.lC, bILine.lMax, bILine.lCov, bILine.fC, bILine.fMax, bILine.fCov = \
                        covLine.lines, covLine.linesMax, covLine.linesCov, covLine.func, covLine.funcMax, covLine.funcCov
                
            ws.auto_filter.ref = "A1:"+get_column_letter(ws.max_column)+str(ws.max_row - 1)
            adjustColumns(ws)

    except IOError as error:
        logging.warning(f"Cannot find/open '{cov_csv_path}'.\n\nError message: '{error}'\n\nCoverage-Sheet will be skipped...\n")


def appendBuildOverviewInfo(ws, buildInfoLines, buildInfoHeader):
    builds_overview_path = os.path.join(path, BUILD_OVERVIEW_FILE_NAME)

    try:
        with open(builds_overview_path) as f:
            reader = csv.reader(f, delimiter=',')
            for idx,row in enumerate(reader):
                if idx == 0:
                    headerLine = HeaderLine(row, ["right","left","right"]+(len(row)-3)*["left"])
                    headerLine.appendRowToWs(ws)
                    buildInfoHeader.row[0:3] = row[0:3]
                    buildInfoHeader.alignment = ["right","left","right"]
                    buildInfoHeader.row += row[3:]
                    buildInfoHeader.alignment += (len(row)-3)*["center"]
                    bILine = getBuildInfoLineFromTag(buildInfoLines, "Total")
                    bILine.optionList = (len(row)-3)*[""]
                else:
                    bOLine = BuildOverviewLine(row)
                    bOLine.appendRowToWs(ws)
                    bILine = getBuildInfoLineFromTag(buildInfoLines, bOLine.tag)
                    bILine.no, bILine.bTime, bILine.optionList = bOLine.no, bOLine.bTime, bOLine.optionList

            ws.auto_filter.ref = "A1:"+get_column_letter(ws.max_column)+str(ws.max_row)
            adjustColumns(ws)
    except IOError as error:
        logging.warning(f"Cannot find/open '{builds_overview_path}'.\n\nError message: '{error}'\n\nBuild-Overview-Sheet will be skipped...\n")
                    

def main(path):
    buildInfoLines = dict()

    res_xlsx_path = os.path.join(path, OUTPUT_FILE_NAME)

    wb = openpyxl.Workbook()
    buildInfoHeader = HeaderLine(10*[""],6*["left"])

    wsCov = wb.active
    wsCov.title = "Coverage"
    appendCoverageInfo(wsCov, buildInfoLines, buildInfoHeader)

    wsBO = wb.create_sheet("Build Overview")
    appendBuildOverviewInfo(wsBO, buildInfoLines, buildInfoHeader)

    wsBA = wb.create_sheet("Build Accesses")
    appendBuildAccessInfo(wsBA, buildInfoLines, buildInfoHeader)

    wsSum = wb.create_sheet("Summary")
    appendBuildInfo(wsSum, buildInfoLines, buildInfoHeader)

    try:
        wb.save(res_xlsx_path)
        logging.info(f"Done! Output file: '{res_xlsx_path}'")
    except IOError as error:
        logging.error(f"Cannot save excel file. Did you close it?\nError message: '{error}'.")


def getHelpString(pName):
    s = "Syntax is:\n"+\
       f"  python3 {pName} <dir_with_info_csv_files>\n\n"+\
        "<dir_with_info_csv_files> needs to contain the following files:\n"+\
        f"'{BUILD_ACCESSES_FILE_NAME}', '{BUILD_OVERVIEW_FILE_NAME}' and '{BUILD_COVERAGE_FILE_NAME}'"

    return s
    

if __name__ == '__main__':
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    
    if len(sys.argv) < 2:
        print("Syntax error. Missing directory path.", file=sys.stderr)
        print(getHelpString(sys.argv[0]), file=sys.stderr)
        exit(1)
    
    path = sys.argv[1]
    if not os.path.isdir(path):
        print(f"Error: Directory '{path}' cannot be found.", file=sys.stderr)
        print(getHelpString(sys.argv[0]), file=sys.stderr)
        exit(1)
    
    main(path)